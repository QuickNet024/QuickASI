# -*- coding: utf-8 -*-
"""WB佣金匹配服务 — 支持多平台多店铺类型、动态多表架构"""

import json
import os
import openpyxl
import logging
from typing import Optional
from datetime import datetime

from src.config import Config
from src.models.database import DatabaseManager
from src.models.commission import Commission, CommissionTableInfo

logger = logging.getLogger(__name__)


class CommissionService:
    def __init__(self, db: DatabaseManager = None):
        self.db = db or DatabaseManager(Config.DB_COMMISSION_PATH, init_tables=["commissions", "commission_meta", "app_config"])

    def import_from_excel(self, file_path: str,
                          platform: str = "wb",
                          shop_type: str = "local") -> int:
        """Import commission table from Excel file.

        Reads ALL columns from the Excel file (not just column C).
        Creates a dynamic table per platform+shop_type combination.

        Args:
            file_path: Excel文件路径
            platform: 平台标识 ('wb', 'ozon', 'market')
            shop_type: 店铺类型 ('local', 'cross_border')
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 1. Read header row (row 1)
        headers = [str(cell or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        # 2. Identify columns:
        #    - First 2 non-empty columns are category and product (always)
        #    - Remaining columns that contain "%" or numeric data are rate columns
        rate_col_indices = []
        rate_headers = []
        for i, h in enumerate(headers):
            if i < 2:
                continue
            if "%" in h or "％" in h or h:
                rate_col_indices.append(i)
                rate_headers.append(h)

        # 3. Build DB column definitions
        rate_db_names = [f"rate_col_{i}" for i in range(len(rate_col_indices))]
        columns = [(name, "REAL") for name in rate_db_names]

        # 4. Create/recreate the data table
        table_name = f"commission_{platform}_{shop_type}"
        self.db.create_commission_data_table(table_name, columns)

        # 5. Read all data rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            category = str(row[0] or "").strip()
            product = str(row[1] or "").strip()
            if not category and not product:
                continue

            rate_values = []
            for idx in rate_col_indices:
                val = row[idx] if idx < len(row) else None
                if val is None:
                    rate_values.append(0.0)
                else:
                    val_str = str(val).replace(",", ".").strip()
                    try:
                        rate_values.append(float(val_str))
                    except ValueError:
                        rate_values.append(0.0)

            rows.append(tuple([category, product] + rate_values))

        wb.close()

        # 6. Batch insert
        if rows:
            self.db.import_commission_rows_batch(table_name, rows)

        # 7. Save metadata
        info = CommissionTableInfo(
            table_name=table_name,
            platform=platform,
            shop_type=shop_type,
            source_file=os.path.basename(file_path),
            column_headers=json.dumps(headers, ensure_ascii=False),
            rate_columns=json.dumps(rate_db_names),
            row_count=len(rows),
            imported_at=datetime.now().isoformat()
        )
        self.db.save_commission_table_info(info)

        return len(rows)

    def sync_from_api(self, platform: str = "wb",
                      shop_type: str = "local") -> int:
        """通过 API 同步佣金数据（预留接口）。

        目前为占位实现，返回 0 并提示暂不支持。
        未来可对接 WB / OZON / MARKET 的佣金 API。
        """
        logger.info(f"Commission API sync requested: {platform}/{shop_type}")
        raise NotImplementedError(
            f"佣金 API 同步暂未开放 ({platform.upper()} "
            f"{'本土' if shop_type == 'local' else '跨境'})。\n"
            f"请使用「导入佣金表」从 Excel 文件导入数据。"
        )

    def find_commission_rate(self, category_name: str,
                             platform: str = "wb",
                             shop_type: str = "local",
                             default_rate: float = None) -> tuple:
        """Find commission rate for a category. Returns (rate, matched, source).

        Looks up from dynamic commission tables first, falls back to
        old commissions table for backward compatibility.
        Default: 30% if no match.
        """
        table_name = f"commission_{platform}_{shop_type}"

        # Check if dynamic table exists
        tables = self.db.get_commission_tables()
        table_names = [t.table_name for t in tables]
        if table_name not in table_names:
            # Fall back to old commissions table for backward compatibility
            rate = self.db.find_commission_by_product(category_name, platform, shop_type)
            if rate is not None:
                return rate, True, "product"
            rate = self.db.find_commission_by_category(category_name, platform, shop_type)
            if rate is not None:
                return rate, True, "category"
            _fallback = default_rate if default_rate is not None else Config.DEFAULT_COMMISSION_RATE
            return _fallback, False, "default"

        # New table lookup — use first rate column
        rate = self.db.find_rate_in_table(table_name, category_name, "rate_col_0")
        if rate is not None:
            return rate, True, "product"
        _fallback = default_rate if default_rate is not None else Config.DEFAULT_COMMISSION_RATE
        return _fallback, False, "default"
