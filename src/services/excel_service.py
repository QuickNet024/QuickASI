import openpyxl
import logging
import shutil
from typing import List, Dict, Optional
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class ProductRow:
    """Represents one row from the uploaded Excel template"""
    row_number: int        # 1-based row number in the sheet
    brand: str            # A
    category: str         # B
    wb_article: str       # C
    seller_sku: str       # D
    barcode: str          # E
    wb_stock: any         # F
    seller_stock: any     # G
    turnover: str         # H
    current_price: str    # I
    new_price: Optional[float]  # J
    current_discount: Optional[int]  # K
    new_discount: Optional[int]      # L (to fill)


class ExcelService:
    # Column indices (1-based)
    COL_SELLER_SKU = 4     # D
    COL_CURRENT_PRICE = 9  # I
    COL_NEW_PRICE = 10     # J
    COL_CURRENT_DISCOUNT = 11  # K
    COL_NEW_DISCOUNT = 12  # L

    def read_template(self, file_path: str) -> List[ProductRow]:
        """Read all product rows from the uploaded Excel template."""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            seller_sku = row[self.COL_SELLER_SKU - 1] if len(row) >= self.COL_SELLER_SKU else None
            if not seller_sku:
                continue

            current_price_raw = row[self.COL_CURRENT_PRICE - 1] if len(row) >= self.COL_CURRENT_PRICE else None
            new_price_raw = row[self.COL_NEW_PRICE - 1] if len(row) >= self.COL_NEW_PRICE else None
            current_discount_raw = row[self.COL_CURRENT_DISCOUNT - 1] if len(row) >= self.COL_CURRENT_DISCOUNT else None
            new_discount_raw = row[self.COL_NEW_DISCOUNT - 1] if len(row) >= self.COL_NEW_DISCOUNT else None

            rows.append(ProductRow(
                row_number=row_idx,
                brand=str(row[0] or "") if len(row) > 0 else "",
                category=str(row[1] or "") if len(row) > 1 else "",
                wb_article=str(row[2] or "") if len(row) > 2 else "",
                seller_sku=str(seller_sku),
                barcode=str(row[4] or "") if len(row) > 4 else "",
                wb_stock=row[5] if len(row) > 5 else None,
                seller_stock=row[6] if len(row) > 6 else None,
                turnover=str(row[7] or "") if len(row) > 7 else "",
                current_price=str(current_price_raw or ""),
                new_price=self._safe_float(new_price_raw),
                current_discount=self._safe_int(current_discount_raw),
                new_discount=self._safe_int(new_discount_raw),
            ))
        wb.close()
        return rows

    def write_updates(self, file_path: str, output_path: str,
                      discount_updates: Dict[int, int] = None,
                      price_updates: Dict[int, float] = None) -> str:
        """Write discount/price updates to a COPY of the Excel file.
        discount_updates/price_updates: {row_number: value}
        Returns the output file path."""
        discount_updates = discount_updates or {}
        price_updates = price_updates or {}

        # Work on a copy
        shutil.copy2(file_path, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        for row_num, discount in discount_updates.items():
            # Clamp: no negatives, max 95
            discount = max(0, min(95, int(discount)))
            cell = ws.cell(row=row_num, column=self.COL_NEW_DISCOUNT)
            cell.value = discount  # Must be int, not string

        for row_num, price in price_updates.items():
            # Clamp: min 5, max 850000
            price = max(5, min(850000, round(float(price), 2)))
            cell = ws.cell(row=row_num, column=self.COL_NEW_PRICE)
            cell.value = price  # Must be float, not string

        wb.save(output_path)
        wb.close()
        return output_path

    def get_seller_skus(self, file_path: str) -> Dict[str, int]:
        """Get all seller SKUs mapped to their row numbers."""
        rows = self.read_template(file_path)
        return {r.seller_sku: r.row_number for r in rows}

    @staticmethod
    def _safe_float(val) -> Optional[float]:
        if val is None:
            return None
        try:
            return float(str(val).replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _safe_int(val) -> Optional[int]:
        if val is None:
            return None
        try:
            return int(float(str(val).replace(",", "").strip()))
        except (ValueError, TypeError):
            return None
