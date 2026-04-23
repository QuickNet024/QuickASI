import pytest
import os
import shutil
import tempfile
import openpyxl
from src.services.excel_service import ExcelService


@pytest.fixture
def test_excel():
    """Create a minimal test Excel file mimicking the WB template structure."""
    tmpdir = tempfile.mkdtemp()
    file_path = os.path.join(tmpdir, "test_template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    # Headers (matching the real template)
    headers = ["品牌", "类目", "WB货号", "卖家货号", "最后条形码", "WB库存", "卖家库存", "周转率", "当前价格", "新价格, CNY", "当前折扣", "新折扣", "折后价格", "存在错误"]
    ws.append(headers)
    # Data rows
    ws.append(["", "音响", 912658, "3C5-YX-JC-530-白-F4", "204992", 0, 220, "", "937", None, 72, None, 262.36, ""])
    ws.append(["", "箱包", 902985, "XB1-13019#-黑色", "204983", 0, 196, "", "563", None, 72, None, 157.64, ""])
    ws.append(["", "电子", 897449, "3C5-YX-TO-T300-橙-F1", "204977", 0, 344, "", "283", None, 73, None, 76.41, ""])
    wb.save(file_path)
    wb.close()
    yield file_path
    shutil.rmtree(tmpdir, ignore_errors=True)


@pytest.fixture
def svc():
    return ExcelService()


class TestReadTemplate:
    def test_reads_all_rows(self, svc, test_excel):
        rows = svc.read_template(test_excel)
        assert len(rows) == 3

    def test_parses_seller_sku(self, svc, test_excel):
        rows = svc.read_template(test_excel)
        assert rows[0].seller_sku == "3C5-YX-JC-530-白-F4"

    def test_parses_category(self, svc, test_excel):
        rows = svc.read_template(test_excel)
        assert rows[0].category == "音响"

    def test_parses_current_price(self, svc, test_excel):
        rows = svc.read_template(test_excel)
        assert rows[0].current_price == "937"

    def test_parses_current_discount(self, svc, test_excel):
        rows = svc.read_template(test_excel)
        assert rows[0].current_discount == 72

    def test_new_price_is_none(self, svc, test_excel):
        rows = svc.read_template(test_excel)
        assert rows[0].new_price is None

    def test_new_discount_is_none(self, svc, test_excel):
        rows = svc.read_template(test_excel)
        assert rows[0].new_discount is None


class TestWriteUpdates:
    def test_write_discount(self, svc, test_excel):
        tmpdir = os.path.dirname(test_excel)
        output = os.path.join(tmpdir, "output.xlsx")
        svc.write_updates(test_excel, output, discount_updates={2: 50})

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        # Column L (12) should have 50
        assert ws.cell(row=2, column=12).value == 50
        # Other cells unchanged
        assert ws.cell(row=2, column=4).value == "3C5-YX-JC-530-白-F4"
        assert ws.cell(row=2, column=11).value == 72
        wb.close()

    def test_write_price(self, svc, test_excel):
        tmpdir = os.path.dirname(test_excel)
        output = os.path.join(tmpdir, "output2.xlsx")
        svc.write_updates(test_excel, output, price_updates={3: 600.0})

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws.cell(row=3, column=10).value == 600.0
        wb.close()

    def test_discount_no_negative(self, svc, test_excel):
        tmpdir = os.path.dirname(test_excel)
        output = os.path.join(tmpdir, "output3.xlsx")
        svc.write_updates(test_excel, output, discount_updates={2: -5})

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=12).value == 0  # Clamped to 0
        wb.close()

    def test_discount_max_95(self, svc, test_excel):
        tmpdir = os.path.dirname(test_excel)
        output = os.path.join(tmpdir, "output4.xlsx")
        svc.write_updates(test_excel, output, discount_updates={2: 99})

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=12).value == 95  # Clamped to 95
        wb.close()

    def test_original_unchanged(self, svc, test_excel):
        """Verify the original file is NOT modified"""
        tmpdir = os.path.dirname(test_excel)
        output = os.path.join(tmpdir, "output5.xlsx")
        svc.write_updates(test_excel, output, discount_updates={2: 50})

        # Original should still have None in column L
        wb = openpyxl.load_workbook(test_excel)
        ws = wb.active
        assert ws.cell(row=2, column=12).value is None
        wb.close()


class TestGetSellerSkus:
    def test_returns_mapping(self, svc, test_excel):
        skus = svc.get_seller_skus(test_excel)
        assert "3C5-YX-JC-530-白-F4" in skus
        assert skus["3C5-YX-JC-530-白-F4"] == 2
